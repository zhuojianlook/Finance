import streamlit as st
import pandas as pd
import re
import openpyxl
from io import BytesIO
import altair as alt

st.title("PI Grant Statement Checker")

# Load sheet names, grant period, and budget total
def load_sheets_and_period(data):
    bio1, bio2 = BytesIO(data), BytesIO(data)
    xls = pd.ExcelFile(bio1, engine='openpyxl')
    default_idx = next((i for i, name in enumerate(xls.sheet_names) if "Details" in name), 0)
    try:
        wb = openpyxl.load_workbook(filename=bio2, data_only=True)
        ws = wb[xls.sheet_names[default_idx]]
        grant_period = ws['B4'].value
        budget_total = float(ws['I8'].value) if ws['I8'].value is not None else None
    except Exception:
        grant_period, budget_total = None, None
    return xls, default_idx, grant_period, budget_total

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"])
if uploaded_file:
    data = uploaded_file.read()
    xls, default_idx, grant_period, budget_total = load_sheets_and_period(data)

    # Sheet selection
    sheet_name = st.selectbox("Select sheet to analyze", xls.sheet_names, index=default_idx)
    if grant_period:
        st.markdown(f"**Grant Period:** {grant_period}")
    if budget_total is not None:
        st.markdown(f"**Budget Amount (cell I8):** {budget_total}")

    # Load data starting row 7, skip row 8
    df = pd.read_excel(BytesIO(data), sheet_name=sheet_name, header=6, skiprows=[7], engine='openpyxl')
    if 'G/L' not in df.columns:
        st.error("Could not find 'G/L' column in sheet.")
        st.stop()
    df = df[df['G/L'].notna()]
    df['Open Items'] = pd.to_numeric(df['Open Items'], errors='coerce')
    df['Invoice'] = pd.to_numeric(df['Invoice'], errors='coerce')

    # Reservations and invoices
    reservations = df[df['Open Items'].notna() & (df['Open Items'] != 0)].copy()
    invoices = df[df['Invoice'].notna() & (df['Invoice'] != 0)].copy()

    # Reservation summary and status
    reservation_summary = (
        reservations.groupby('Text')['Open Items'].sum()
        .reset_index().rename(columns={'Open Items': 'Net Reservation Amount'})
    )
    res_out = reservation_summary[reservation_summary['Net Reservation Amount'] != 0].copy()
    invoices['InvAmtRnd'] = invoices['Invoice'].round(2)
    invoice_counts = (invoices['InvAmtRnd'].value_counts().rename_axis('InvAmtRnd')
                      .reset_index(name='Invoice Count'))
    res_out['InvAmtRnd'] = res_out['Net Reservation Amount'].round(2)
    res_out = res_out.merge(invoice_counts, how='left', on='InvAmtRnd')
    res_out['Invoice Count'] = res_out['Invoice Count'].fillna(0).astype(int)
    amt_counts = (res_out.groupby('InvAmtRnd')['Text'].count()
                  .rename('Reservation Codes Count').reset_index())
    res_out = res_out.merge(amt_counts, on='InvAmtRnd')
    def status(r):
        ic, rc = r['Invoice Count'], r['Reservation Codes Count']
        if r['Net Reservation Amount'] != 0 and ic > 0: return 'Invoiced without Balancing Open Items'
        if ic == 0: return 'Uninvoiced'
        if ic > rc: return 'Multiple invoices'
        if ic == rc == 1: return 'Invoiced'
        if ic == 1 and rc > 1: return 'Ambiguous reservation codes'
        return 'Check'
    res_out['Status'] = res_out.apply(status, axis=1)

    # 1️⃣ Pending Invoices and Errors
    st.subheader("Pending Invoices and Errors")
    st.dataframe(res_out[['Text','Net Reservation Amount','Invoice Count','Reservation Codes Count','Status']])

    # 2️⃣ Duplicated Invoice Charges
    st.subheader("Duplicated Invoice Charges")
    dup = invoices[invoices.duplicated(subset=['InvAmtRnd','Text'],keep=False)]
    st.dataframe(dup[['Posting Date','Ref Document','Text','Invoice']])

    # 3️⃣ Ambiguous Invoices
    st.subheader("Ambiguous Invoices (matched by $ amount)")
    def extract_R(codes): return [c for c in codes if re.match(r'^R\d+', c)]
    pos_map = (reservations[reservations['Open Items']>0]
               .assign(OpenRnd=lambda d: d['Open Items'].round(2))
               .groupby('OpenRnd')['Text'].apply(lambda x:list(set(x))).to_dict())
    amb = invoices.copy()
    amb['Possible R-Codes'] = amb['InvAmtRnd'].map(pos_map).apply(lambda lst: extract_R(lst) if isinstance(lst,list) else [])
    amb_flag = amb[amb['Possible R-Codes'].apply(len)>1]
    st.dataframe(amb_flag[['Posting Date','Ref Document','Text','Invoice','Possible R-Codes']])

    # 4️⃣ All Invoices
    st.subheader("All Invoices")
    st.dataframe(invoices[['Posting Date','Ref Document','Text','Invoice']])

    # 5️⃣ Remaining Budget over Time
    if budget_total and grant_period:
        try:
            start_str, end_str = [s.strip() for s in grant_period.split(' to ')]
            start_date = pd.to_datetime(start_str, format='%d.%m.%Y', dayfirst=True)
            end_date = pd.to_datetime(end_str, format='%d.%m.%Y', dayfirst=True)
        except:
            start_date, end_date = None, None

        if start_date and end_date:
            spend = (invoices.assign(Date=pd.to_datetime(invoices['Posting Date'], dayfirst=True, errors='coerce'))
                     .groupby('Date')['Invoice'].sum().cumsum().rename('Cumulative Spend'))
            # Obligations events for uninvoiced items at their latest reservation date
            uninv = res_out[res_out['Status']=='Uninvoiced'][['Text','Net Reservation Amount']]
            latest = (
                reservations.groupby('Text')['Posting Date']
                .max()
                .reset_index()
                .rename(columns={'Posting Date':'Date'})
            )
            # Ensure 'Date' is datetime
            latest['Date'] = pd.to_datetime(latest['Date'], dayfirst=True, errors='coerce')
            events = pd.merge(uninv, latest, on='Text', how='inner').rename(columns={'Net Reservation Amount':'Amount'})
            obligations = events.groupby('Date')['Amount'].sum().cumsum().rename('Obligations')
            # Create timeline index correctly
            idx = pd.date_range(start_date, end_date)
            spend_df = spend.reindex(idx, method='ffill').fillna(0)
            obligations_df = obligations.reindex(idx, method='ffill').fillna(0)
            remaining = budget_total - spend_df
            remaining_alloc = budget_total - (spend_df + obligations_df)
            timeline = pd.DataFrame({'Remaining Budget':remaining,
                                     'Remaining (incl. Allocated)':remaining_alloc},
                                     index=idx).reset_index().rename(columns={'index':'Date'})
            y_scale = alt.Scale(domain=[0,budget_total])
            today = pd.Timestamp.today().normalize()
            days_left = max(0,(end_date-today).days)
            total_days = (end_date-start_date).days
            pct_time_left = days_left/total_days*100 if total_days else None
            total_used = spend_df.iloc[-1]
            pct_used = total_used/budget_total*100 if budget_total else None
            st.markdown(f"**Time Left:** {days_left} days ({pct_time_left:.1f}% of period)")
            st.markdown(f"**Total Utilized:** {total_used} ({pct_used:.1f}% utilized)")
            total_alloc_util = total_used + obligations_df.iloc[-1]
            pct_alloc_util = total_alloc_util/budget_total*100 if budget_total else None
            st.markdown(f"**Total Allocated + Utilized:** {total_alloc_util} ({pct_alloc_util:.1f}% of budget)")
            st.subheader("Budget Usage over Time")
            melt = timeline.melt('Date',var_name='Metric',value_name='Value')
            chart = alt.Chart(melt).mark_line().encode(x=alt.X('Date:T',title='Date'),
                                                     y=alt.Y('Value:Q',title='Amount',scale=y_scale),
                                                     color=alt.Color('Metric:N',title='Series')).properties(width=700,height=400)
            expiry = alt.Chart(pd.DataFrame({'Date':[end_date]})).mark_rule(color='red').encode(x='Date:T')
            today_rule = alt.Chart(pd.DataFrame({'Date':[today]})).mark_rule(color='gray',strokeDash=[4,4]).encode(x='Date:T')
            st.altair_chart(chart+expiry+today_rule, use_container_width=True)
